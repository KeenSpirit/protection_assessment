"""
Legacy SEQ fault level study Excel output generation.

This module generates Excel workbooks in the original SEQ Protection
department format. It creates multi-sheet workbooks with input summaries,
fault level results, and detailed per-terminal data.

Output Structure:
    - Inputs Summary: Study parameters, external grid data
    - Results Summary: Per-device fault levels and open points
    - {Feeder} sheets: Detailed terminal and line-level results

Save Locations (tried in order):
    1. Y:/PROTECTION/PowerFactory/Fault Level Studies (department drive)
    2. //client/{home_drive}$/{home_path} (Citrix mapping)
    3. User home directory (local install)

Functions:
    output_results: Generate Excel workbook from result dictionaries
    save_results: Save workbook to disk
"""

import datetime
import math
import os
from pathlib import Path
import time
from typing import Any, Dict

from pf_config import pft


def output_results(
    app: pft.Application,
    sub_name: str,
    external_grid: Dict,
    feeders_devices_inrush: Dict,
    results_max_3p: Dict,
    results_max_2p: Dict,
    results_max_pg: Dict,
    results_min_2p: Dict,
    results_min_3p: Dict,
    results_min_pg: Dict,
    result_sys_norm_min_2p: Dict,
    result_sys_norm_min_pg: Dict,
    feeders_sections_trmax_size: Dict,
    results_max_tr_3p: Dict,
    results_max_tr_pg: Dict,
    results_all_max_3p: Dict,
    results_all_max_2p: Dict,
    results_all_max_pg: Dict,
    results_all_min_3p: Dict,
    results_all_min_2p: Dict,
    results_all_min_pg: Dict,
    result_all_sys_norm_min_2p: Dict,
    result_all_sys_norm_min_pg: Dict,
    feeders_devices_load: Dict,
    results_lines_max_3p: Dict,
    results_lines_max_2p: Dict,
    results_lines_max_pg: Dict,
    results_lines_min_3p: Dict,
    results_lines_min_2p: Dict,
    results_lines_min_pg: Dict,
    result_lines_sys_norm_min_2p: Dict,
    result_lines_sys_norm_min_pg: Dict,
    result_lines_type: Dict,
    result_lines_therm_rating: Dict,
    fdrs_open_switches: Dict
) -> Any:
    """
    Generate legacy format Excel workbook from result dictionaries.

    Creates a multi-sheet workbook with fault study results in the
    original SEQ Protection department format.

    Args:
        app: PowerFactory application instance.
        sub_name: Substation name string for titles.
        external_grid: External grid parameters dictionary.
        feeders_devices_inrush: Inrush current per device.
        results_max_3p: Maximum 3-phase fault at device.
        results_max_2p: Maximum 2-phase fault at device.
        results_max_pg: Maximum phase-ground fault at device.
        results_min_2p: Minimum 2-phase fault at device.
        results_min_3p: Minimum 3-phase fault at device.
        results_min_pg: Minimum phase-ground fault at device.
        result_sys_norm_min_2p: System normal minimum 2-phase fault.
        result_sys_norm_min_pg: System normal minimum P-G fault.
        feeders_sections_trmax_size: Largest transformer per section.
        results_max_tr_3p: Max 3-phase at largest transformer.
        results_max_tr_pg: Max P-G at largest transformer.
        results_all_max_3p: Max 3-phase at all terminals.
        results_all_max_2p: Max 2-phase at all terminals.
        results_all_max_pg: Max P-G at all terminals.
        results_all_min_3p: Min 3-phase at all terminals.
        results_all_min_2p: Min 2-phase at all terminals.
        results_all_min_pg: Min P-G at all terminals.
        result_all_sys_norm_min_2p: System normal min 2-phase at terminals.
        result_all_sys_norm_min_pg: System normal min P-G at terminals.
        feeders_devices_load: Transformer loads per section.
        results_lines_max_3p: Max 3-phase at all lines.
        results_lines_max_2p: Max 2-phase at all lines.
        results_lines_max_pg: Max P-G at all lines.
        results_lines_min_3p: Min 3-phase at all lines.
        results_lines_min_2p: Min 2-phase at all lines.
        results_lines_min_pg: Min P-G at all lines.
        result_lines_sys_norm_min_2p: System normal min 2-phase at lines.
        result_lines_sys_norm_min_pg: System normal min P-G at lines.
        result_lines_type: Conductor type per line.
        result_lines_therm_rating: Thermal rating per line.
        fdrs_open_switches: Open switches per feeder.

    Returns:
        openpyxl Workbook object ready for saving.
    """
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    app.PrintPlain("Creating output file...")

    wb = openpyxl.Workbook()
    heading_font = Font(size=14, bold=True)
    subheading_font = Font(size=11, bold=True)

    # Create Inputs Summary sheet
    _create_inputs_sheet(
        wb, sub_name, external_grid, heading_font, subheading_font
    )

    # Create Results Summary sheet
    _create_results_sheet(
        wb, sub_name, feeders_devices_inrush, heading_font
    )

    # Create per-feeder detailed sheets
    _create_feeder_sheets(
        wb, sub_name, feeders_devices_inrush, heading_font
    )

    # Write data to sheets
    r_s = wb["Results Summary"]

    # Write nested dictionary data using helper functions
    _nested_dic(r_s, feeders_devices_inrush, 1, "Inrush (A):", heading_font)
    _two_nested_dic(r_s, results_max_3p, 2, "Max 3-P fault level (kA) (Site):")
    _two_nested_dic(r_s, results_max_2p, 3, "Max 2-P fault level (kA) (Site):")
    _two_nested_dic(r_s, results_max_pg, 4, "Max P-G fault level (kA) (Site):")
    _two_nested_dic(r_s, results_min_3p, 5, "Min 3-P fault level (kA) (Site):")
    _two_nested_dic(r_s, results_min_2p, 6, "Min 2-P fault level (kA) (Site):")
    _two_nested_dic(r_s, results_min_pg, 7, "Min P-G fault level (kA) (Site):")
    _two_nested_dic(
        r_s, result_sys_norm_min_2p, 8,
        "System normal Min 2-P fault level (kA) (Site):"
    )
    _two_nested_dic(
        r_s, result_sys_norm_min_pg, 9,
        "System normal Min P-G fault level (kA) (Site):"
    )
    _two_nested_dic(
        r_s, feeders_sections_trmax_size, 10,
        "Largest transformer size (kVA) (Site):"
    )
    _two_nested_dic(
        r_s, results_max_tr_3p, 11,
        "Max 3-P at largest transformer (kA) (Site):"
    )
    _two_nested_dic(
        r_s, results_max_tr_pg, 12,
        "Max P-G at largest transformer (kA) (Site):"
    )

    _elements_open(
        r_s, fdrs_open_switches, feeders_devices_inrush, heading_font
    )

    # Write detailed sheets
    _sheets_nested_dic(
        wb, feeders_devices_load, result_lines_type,
        -2, -2, "Tfmr load (kVA)", "Conductor type", heading_font
    )
    _sheets_nested_dic(
        wb, results_all_max_3p, result_lines_therm_rating,
        0, 0, "Max 3P fault", "Rated 1s current (kA)", heading_font
    )
    _sheets_nested_dic(
        wb, results_all_max_2p, results_lines_max_3p,
        1, 1, "Max 2P fault", "Max 3P fault", heading_font
    )
    _sheets_nested_dic(
        wb, results_all_max_pg, results_lines_max_2p,
        2, 2, "Max PG fault", "Max 2P fault", heading_font
    )
    _sheets_nested_dic(
        wb, results_all_min_3p, results_lines_max_pg,
        3, 3, "Min 3P fault", "Max PG fault", heading_font
    )
    _sheets_nested_dic(
        wb, results_all_min_2p, results_lines_min_3p,
        4, 4, "Min 2P fault", "Min 3P fault", heading_font
    )
    _sheets_nested_dic(
        wb, results_all_min_pg, results_lines_min_2p,
        5, 5, "Min PG fault", "Min 2P fault", heading_font
    )
    _sheets_nested_dic(
        wb, result_all_sys_norm_min_2p, results_lines_min_pg,
        6, 6, "Min Sys Norm 2P fault", "Min PG fault", heading_font
    )
    _sheets_nested_dic(
        wb, result_all_sys_norm_min_pg, result_lines_sys_norm_min_2p,
        7, 7, "Min Sys Norm PG fault", "Min Sys Norm 2P fault", heading_font
    )
    _sheets_nested_dic(
        wb, results_all_max_3p, result_lines_sys_norm_min_pg,
        0, 8, "Max 3P fault", "Min Sys Norm PG fault", heading_font
    )

    return wb


def _create_inputs_sheet(
    wb: Any,
    sub_name: str,
    external_grid: Dict,
    heading_font: Any,
    subheading_font: Any
) -> None:
    """
    Create and populate the Inputs Summary sheet.

    Args:
        wb: openpyxl Workbook.
        sub_name: Substation name.
        external_grid: External grid parameters.
        heading_font: Font for headings.
        subheading_font: Font for subheadings.
    """
    wb.create_sheet(index=0, title="Inputs Summary")
    wb.remove(wb["Sheet"])
    i_s = wb["Inputs Summary"]

    # Column widths
    i_s.column_dimensions['A'].width = 19.86
    i_s.column_dimensions['B'].width = 9.14
    i_s.column_dimensions['C'].width = 9.14

    # Fonts
    i_s['A1'].font = heading_font
    i_s['A3'].font = subheading_font
    i_s['A7'].font = heading_font
    i_s['A13'].font = subheading_font
    i_s['A22'].font = subheading_font

    # Static content
    i_s['A1'] = sub_name + " Feeder Fault Level Study"
    i_s['A3'] = "Script Run Date:"
    i_s['A4'] = datetime.datetime.now()
    i_s['A7'] = "Input Summary"
    i_s['A9'] = "Short-circuit calculation method: Complete"
    i_s['A10'] = "c-Factor (max): 1.1"
    i_s['C10'] = "Voltage factor c (max): 1.1"
    i_s['A11'] = "c-Factor (min): 1.0"
    i_s['C11'] = "Voltage factor c (min): 1.0"
    i_s['A13'] = "External Grid Data:"
    i_s['A16'] = "3-P fault level (A):"
    i_s['A17'] = "R/X:"
    i_s['A18'] = "Z2/Z1:"
    i_s['A19'] = "X0/X1:"
    i_s['A20'] = "R0/X0:"
    i_s['A22'] = (
        "NOTE: Open points to adjacent bulk supply substations may not be "
        "detected unless the relevant grids are active under the current "
        "project"
    )

    # External grid data
    next_column = 2
    for grid, value in external_grid.items():
        i_s.cell(
            column=next_column, row=14, value=grid.loc_name
        ).font = subheading_font
        i_s.cell(column=next_column, row=15, value="Maximum")
        i_s.cell(column=next_column + 1, row=15, value="Minimum")
        i_s.cell(column=next_column + 2, row=15, value="System Normal Minimum")

        # Max values
        next_row = 16
        for values in value[:5]:
            i_s.cell(column=next_column, row=next_row, value=values)
            next_row += 1

        # Min values
        next_row = 16
        for values in value[5:10]:
            i_s.cell(column=next_column + 1, row=next_row, value=values)
            next_row += 1

        # System normal min values
        next_row = 16
        for values in value[10:]:
            i_s.cell(column=next_column + 2, row=next_row, value=values)
            next_row += 1

        next_column += 4


def _create_results_sheet(
    wb: Any,
    sub_name: str,
    feeders_devices_inrush: Dict,
    heading_font: Any
) -> None:
    """
    Create the Results Summary sheet structure.

    Args:
        wb: openpyxl Workbook.
        sub_name: Substation name.
        feeders_devices_inrush: Feeder data for column sizing.
        heading_font: Font for headings.
    """
    from openpyxl.utils import get_column_letter

    wb.create_sheet(index=1, title="Results Summary")
    r_s = wb["Results Summary"]

    # Set column widths
    length = len(feeders_devices_inrush)
    j = 1
    for i in range(length):
        k = get_column_letter(j)
        r_s.column_dimensions[k].width = 37
        j += 5

    r_s.column_dimensions['A'].width = 36.86
    r_s['A1'].font = heading_font
    r_s['A3'].font = heading_font
    r_s['A5'].font = heading_font

    r_s['A1'] = sub_name + " Feeder Fault Level Study"
    r_s['A3'] = "Results Summary"


def _create_feeder_sheets(
    wb: Any,
    sub_name: str,
    feeders_devices_inrush: Dict,
    heading_font: Any
) -> None:
    """
    Create per-feeder detailed result sheets.

    Args:
        wb: openpyxl Workbook.
        sub_name: Substation name.
        feeders_devices_inrush: Feeder names for sheet creation.
        heading_font: Font for headings.
    """
    for key in feeders_devices_inrush.keys():
        wb.create_sheet(title=key)
        feed = wb[key]
        title = f"{sub_name} Feeder Fault Level Study - Detailed Results for {key}"
        feed['A1'] = title
        feed['A1'].font = heading_font


def _nested_dic(
    r_s: Any,
    data: Dict,
    row_offset: int,
    label: str,
    heading_font: Any
) -> None:
    """
    Write single-level nested dictionary data to Results Summary.

    Args:
        r_s: Results Summary worksheet.
        data: Dictionary of feeder -> device -> value.
        row_offset: Row offset within device section.
        label: Row label text.
        heading_font: Font for headings.
    """
    next_column = 2

    for key, value in data.items():
        r_s.cell(
            column=next_column - 1, row=5, value="FEEDER:"
        ).font = heading_font
        r_s.cell(column=next_column, row=5, value=key).font = heading_font

        next_row = 7
        for key2, values2 in value.items():
            r_s.cell(
                column=next_column - 1, row=next_row, value="SECTION:"
            ).font = heading_font
            r_s.cell(
                column=next_column, row=next_row, value=key2
            ).font = heading_font
            r_s.cell(column=next_column - 1, row=next_row + row_offset, value=label)
            r_s.cell(column=next_column, row=next_row + row_offset, value=values2)
            next_row += 14

        next_column += 5


def _two_nested_dic(
    r_s: Any,
    data: Dict,
    row_offset: int,
    label: str
) -> None:
    """
    Write two-level nested dictionary data to Results Summary.

    Args:
        r_s: Results Summary worksheet.
        data: Dictionary of feeder -> device -> terminal -> value.
        row_offset: Row offset within device section.
        label: Row label text.
    """
    next_column = 2

    for key, value in data.items():
        next_row = 7

        for key2, value2 in value.items():
            for key3, value3 in value2.items():
                # Strip '_Term' suffix if present
                key4 = key3[:-5] if key3.endswith("_Term") else key3
                r_s.cell(column=next_column + 1, row=next_row + row_offset, value=key4)
                r_s.cell(
                    column=next_column - 1, row=next_row + row_offset, value=label
                )
                r_s.cell(column=next_column, row=next_row + row_offset, value=value3)

            next_row += 14

        next_column += 5


def _elements_open(
    r_s: Any,
    fdrs_open_switches: Dict,
    feeders_devices_inrush: Dict,
    heading_font: Any
) -> None:
    """
    Write open switch information to Results Summary.

    Args:
        r_s: Results Summary worksheet.
        fdrs_open_switches: Open switches per feeder.
        feeders_devices_inrush: Feeder data for row calculation.
        heading_font: Font for headings.
    """
    feeder_row_dic = {
        feeder: len(feeders_devices_inrush[feeder])
        for feeder in feeders_devices_inrush
    }

    next_column = 2

    for feeder, open_switches in fdrs_open_switches.items():
        last_row = 7 + (feeder_row_dic[feeder] * 14)
        r_s.cell(
            column=next_column - 1, row=last_row, value="FEEDER OPEN POINTS:"
        ).font = heading_font

        if len(open_switches) > 0:
            for site, switch in open_switches.items():
                switch_name = switch.GetAttribute("loc_name")

                if switch.GetClassName() == "StaSwitch":
                    r_s.cell(column=next_column, row=last_row, value=switch_name)
                else:
                    site_name = site.GetAttribute("loc_name")
                    r_s.cell(
                        column=next_column, row=last_row,
                        value=f"{site_name} / {switch_name}"
                    )
                last_row += 1
        else:
            r_s.cell(
                column=next_column - 1, row=last_row, value="(None detected)"
            )
            last_row += 1

        last_row += 1
        next_column += 5


def _sheets_nested_dic(
    wb: Any,
    node_data: Dict,
    line_data: Dict,
    col_node_data: int,
    col_line_data: int,
    col_name_1: str,
    col_name_2: str,
    heading_font: Any
) -> None:
    """
    Write detailed terminal and line data to per-feeder sheets.

    Args:
        wb: openpyxl Workbook.
        node_data: Terminal-level data dictionary.
        line_data: Line-level data dictionary.
        col_node_data: Column offset for node data.
        col_line_data: Column offset for line data.
        col_name_1: Header for node data column.
        col_name_2: Header for line data column.
        heading_font: Font for headings.
    """
    for key, value in node_data.items():
        feed = wb[key]

        # Set column widths
        feed.column_dimensions['B'].width = 18.14
        feed.column_dimensions['M'].width = 18.14
        feed.column_dimensions['W'].width = 18.14
        feed.column_dimensions['AG'].width = 18.14
        feed.column_dimensions['AQ'].width = 18.14

        next_column = 3

        for key2, value2 in value.items():
            feed.cell(
                column=next_column, row=3, value=key2
            ).font = heading_font
            feed.cell(
                column=next_column - 2, row=3, value="SECTION:"
            ).font = heading_font
            feed.cell(column=next_column - 1, row=4, value="Downstream terminal")
            feed.cell(column=next_column + col_node_data, row=4, value=col_name_1)

            next_row = 3

            # Write terminal data
            for key3, value3 in value2.items():
                key4 = key3[:-5] if str(key3).endswith("_Term") else key3
                feed.cell(column=next_column - 1, row=next_row + 2, value=key4)
                feed.cell(
                    column=next_column + col_node_data,
                    row=next_row + 2, value=value3
                )
                next_row += 1

            # Write line section header
            feed.cell(
                column=next_column - 2, row=next_row + 3, value="LINES"
            ).font = heading_font
            feed.cell(column=next_column - 1, row=next_row + 4, value="Line")
            feed.cell(
                column=next_column + col_line_data,
                row=next_row + 4, value=col_name_2
            )

            # Write line data
            length = 0
            line_dict = line_data[key][key2]

            for key5, value5 in line_dict.items():
                feed.cell(
                    column=next_column - 1,
                    row=next_row + length + 5, value=key5.loc_name
                )
                feed.cell(
                    column=next_column + col_line_data,
                    row=next_row + length + 5, value=value5
                )
                length += 1

            next_column += 11


def save_results(
    app: pft.Application,
    sub_name: str,
    wb: Any
) -> None:
    """
    Save Excel workbook to disk.

    Attempts to save to multiple locations in order of preference:
    1. SEQ Protection department shared drive
    2. Citrix client mapping to home directory
    3. Local home directory

    Args:
        app: PowerFactory application instance.
        sub_name: Substation name for filename.
        wb: openpyxl Workbook to save.

    Side Effects:
        Saves workbook to disk and prints location to output window.
    """
    date_string = time.strftime("%Y%m%d-%H%M%S")
    file_name = sub_name + " fault level study " + date_string + ".xlsx"

    home_path = str(Path.home())
    save_path = r'\\client\Y$\PROTECTION\PowerFactory\Fault Level Studies'
    save_path_2 = '\\\\client\\' + home_path[0] + '$' + home_path[2:]
    save_path_3 = home_path

    # Try department drive first
    try:
        wb.save(os.path.join(save_path, file_name))
        app.PrintPlain("Output file saved to " + save_path)
    except Exception:
        # Try Citrix mapping
        try:
            wb.save(os.path.join(save_path_2, file_name))
            app.PrintPlain("Output file saved to " + save_path_2)
        except Exception:
            # Fall back to local home directory
            wb.save(os.path.join(save_path_3, file_name))
            app.PrintPlain("Output file saved to " + save_path_3)