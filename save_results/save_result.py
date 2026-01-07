"""
Excel workbook generation for protection assessment results.

This module creates formatted Excel workbooks containing fault study
results, protection reach factors, and conductor damage assessments.
Output files are saved to the user's local directory.

Output Sheets:
    - General Information: Study parameters, grid data, settings
    - Summary Results: Device fault levels, capacity, backup devices
    - {Feeder} Detailed Results: Terminal-by-terminal analysis
    - {Feeder} Cond Dmg Res: Conductor damage results (if selected)

Functions:
    save_dataframe: Main entry point for Excel output generation
    format_grid_data: Format external grid parameters
    format_fl_results: Format fault level results per feeder
    format_study_results: Format device summary data
    format_fdr_open_points: Format feeder open point list
    format_detailed_results: Format terminal-level results
    format_devices: Create device data structure template
"""

from pathlib import Path
import os
import re
import time
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from pf_config import pft
from fault_study import fault_impedance
from relays.reach_factors import device_reach_factors
from save_results import cond_dmg_results as cd
from importlib import reload

reload(fault_impedance)
reload(cd)


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def save_dataframe(
    app: pft.Application,
    region: str,
    study_selections: List[str],
    external_grid: Dict,
    feeders: List
) -> None:
    """
    Save protection assessment results to an Excel workbook.

    Creates a formatted Excel file containing all study results.
    The file is saved to the user's local directory with a timestamp.

    Args:
        app: PowerFactory application instance.
        region: Network region ('SEQ' or 'Regional Models').
        study_selections: List of selected study types.
        external_grid: Dictionary of grid objects to parameter lists.
        feeders: List of Feeder dataclasses with populated results.

    Output Location:
        Attempts paths in order:
        1. //client/c$/LocalData/{username}/ (Citrix)
        2. c:/LocalData/{username}/ (Local)

    Example:
        >>> save_dataframe(app, 'SEQ', selections, grids, feeders)
        Output file saved to //client/c$/LocalData/user/Fault Study...
    """
    project = app.GetActiveProject()
    derived_proj = project.der_baseproject

    try:
        der_proj_name = derived_proj.GetFullName()
    except AttributeError:
        der_proj_name = project.loc_name

    try:
        project_version = project.der_baseversion
    except AttributeError:
        project_version = 'NA'

    app.PrintPlain("Saving Fault Level Study...")

    # Generate filename with timestamp
    date_string = time.strftime("%Y%m%d-%H%M%S")
    study_case_name = app.GetActiveStudyCase().loc_name
    filename = f'Fault Study Results {study_case_name} {date_string}.xlsx'
    filename = fix_string(filename)

    # Determine output path
    user = Path.home().name
    basepath = Path('//client/c$/LocalData') / user

    if basepath.exists():
        clientpath = basepath
    else:
        clientpath = Path('c:/LocalData') / user

    filepath = os.path.join(clientpath, filename)

    # Format data for output
    formatted_grid_data = format_grid_data(external_grid)
    grid_data_df = pd.DataFrame(formatted_grid_data)
    grid_data_df = clean_dataframe(grid_data_df)

    fault_studies_pd = format_fl_results(app, region, feeders)

    # Regional fault impedance values
    if region == 'SEQ':
        oh_z = '0'
        ug_z = '0'
    else:
        oh_z = '50'
        ug_z = '10'

    variations = app.GetActiveNetworkVariations()

    # Write to Excel
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        workbook = writer.book

        # General Information sheet
        grid_data_df.to_excel(
            writer,
            sheet_name='General Information',
            startrow=26,
            index=False
        )
        worksheet = workbook['General Information']

        _write_general_info(
            worksheet, study_case_name, date_string, der_proj_name,
            project_version, oh_z, ug_z, variations
        )

        # Summary Results and Detailed Results sheets
        i = 1
        for feeder, key in fault_studies_pd.items():
            study_results = clean_dataframe(key[0])
            fdr_open_points = clean_dataframe(key[1])
            dfls_list = key[2]

            # Write summary results
            col_letter = get_column_letter(i)
            study_results.to_excel(
                writer,
                sheet_name='Summary Results',
                startrow=2,
                startcol=i - 1,
                index=False
            )
            fdr_open_points.to_excel(
                writer,
                sheet_name='Summary Results',
                startrow=21,
                startcol=i - 1,
                index=False
            )

            sheet = workbook['Summary Results']
            sheet[f'{col_letter}1'].font = Font(size=12, bold=True)
            safe_set_cell(sheet, f'{col_letter}1', fix_string(str(feeder)))

            study_results_len = len(study_results.columns)
            i = i + study_results_len + 1

            # Write detailed results
            safe_feeder_name = create_safe_sheet_name(
                f"{str(feeder)} Detailed Results"
            )

            for j, device in enumerate(dfls_list):
                count = (j + 1) * 27 - 27
                device_name = (
                    str(device.columns[1])
                    if len(device.columns) > 1
                    else "Unknown Device"
                )

                device = clean_dataframe(device)
                device = ensure_numeric_types(device)

                device.to_excel(
                    writer,
                    sheet_name=safe_feeder_name,
                    startrow=1,
                    startcol=count,
                    index=False
                )

                sheet = workbook[safe_feeder_name]
                start_col = get_column_letter(count + 1)
                sheet[f"{start_col}1"].font = Font(size=11, bold=True)
                safe_set_cell(
                    sheet,
                    f"{start_col}1",
                    f'Primary protection: {device_name}'
                )

            # Conductor Damage Results
            if 'Conductor Damage Assessment' in study_selections:
                devices = [
                    fdr.devices
                    for fdr in feeders
                    if fdr.obj.loc_name == feeder
                ][0]
                cond_damage_df = cd.cond_damage_results(devices)
                cond_damage_df = clean_dataframe(cond_damage_df)
                cond_damage_df = ensure_numeric_types(cond_damage_df)

                sheet_name = create_safe_sheet_name(f"{feeder} Cond Dmg Res")
                cond_damage_df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    startrow=0,
                    index=False
                )

    # Apply formatting
    wb = load_workbook(filepath)
    ws = wb['General Information']
    adjust_gen_info_col_size(ws)

    ws = wb['Summary Results']
    adjust_summ_col_size(ws)

    for feeder in fault_studies_pd:
        safe_name = create_safe_sheet_name(f"{str(feeder)} Detailed Results")
        if safe_name in wb.sheetnames:
            ws = wb[safe_name]
            adjust_detailed_col_size(ws)

            if 'Conductor Damage Assessment' in study_selections:
                sheet_name = create_safe_sheet_name(f"{feeder} Cond Dmg Res")
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    adjust_cond_damage_col_width(ws)

    wb.save(filepath)
    app.PrintPlain("Output file saved to " + filepath)


def _write_general_info(
    worksheet,
    study_case: str,
    date_string: str,
    project_name: str,
    version: str,
    oh_z: str,
    ug_z: str,
    variations: List
) -> None:
    """Write general information content to worksheet."""
    safe_set_cell(worksheet, 'A1', str(study_case))
    safe_set_cell(worksheet, 'A2', "Fault Level Study")
    safe_set_cell(worksheet, 'A4', 'Script Run Date-Time')
    safe_set_cell(worksheet, 'A5', str(date_string))
    safe_set_cell(worksheet, 'A6', 'Base Project:')
    safe_set_cell(worksheet, 'A7', str(project_name))
    safe_set_cell(worksheet, 'A8', 'Used Version:')
    safe_set_cell(worksheet, 'A9', str(version))
    safe_set_cell(worksheet, 'A10', 'Network Variations:')

    if variations:
        var_names = ', '.join([v.loc_name for v in variations])
        safe_set_cell(worksheet, 'A11', var_names)
    else:
        safe_set_cell(worksheet, 'A11', 'None')

    safe_set_cell(worksheet, 'A13', 'Short-circuit calculation method:')
    safe_set_cell(worksheet, 'B13', 'Complete')
    safe_set_cell(worksheet, 'A14', 'Maximum fault c-factor:')
    safe_set_cell(worksheet, 'B14', '1.1')
    safe_set_cell(worksheet, 'A15', 'Minimum fault c-factor:')
    safe_set_cell(worksheet, 'B15', '1.0')
    safe_set_cell(worksheet, 'A17', 'OH Line Minimum Earth Fault Impedance:')
    safe_set_cell(worksheet, 'B17', f'{oh_z} ohms')
    safe_set_cell(worksheet, 'A18', 'UG Cable Minimum Earth Fault Impedance:')
    safe_set_cell(worksheet, 'B18', f'{ug_z} ohms')
    safe_set_cell(worksheet, 'A20', 'Reach Factor Thresholds:')
    safe_set_cell(worksheet, 'A21', 'Primary RF >= 2.0 (SEQ) or 1.7 (Regional)')
    safe_set_cell(worksheet, 'A22', 'Backup RF >= 1.3')
    safe_set_cell(
        worksheet,
        'A23',
        'Tabulated fault clearing time shown is for final trip.'
    )
    safe_set_cell(worksheet, 'A25', 'External Grid Data:')


# =============================================================================
# DATA FORMATTING FUNCTIONS
# =============================================================================

def format_grid_data(ext_grid: Dict) -> Dict:
    """
    Format external grid parameters for DataFrame creation.

    Args:
        ext_grid: Dictionary mapping grid objects to parameter lists.

    Returns:
        Dictionary with 'Parameter' column and columns for each grid's
        Maximum, Minimum, and System Normal Minimum values.
    """
    formatted_grid_data = {}

    for grid, attributes in ext_grid.items():
        formatted_grid_data['Parameter'] = [
            '3-P fault level (A):',
            'R/X:',
            'Z2/Z1:',
            'X0/X1:',
            'R0/X0:'
        ]
        formatted_grid_data[f'{grid.loc_name} Maximum'] = attributes[:5]
        formatted_grid_data[f'{grid.loc_name} Minimum'] = attributes[5:10]
        formatted_grid_data[f'{grid.loc_name} Sys Norm Minimum'] = attributes[-5:]

    return formatted_grid_data


def format_fl_results(app: pft.Application, region: str, feeders: List) -> Dict:
    """
    Format fault level results for all feeders.

    Args:
        app: PowerFactory application instance.
        region: Network region string.
        feeders: List of Feeder dataclasses.

    Returns:
        Dictionary mapping feeder names to lists containing:
        [study_results_df, open_points_df, detailed_results_list]
    """
    fault_studies_pd = {}

    for feeder in feeders:
        study_results_df = format_study_results(app, feeder.devices)
        fdr_open_points = format_fdr_open_points(feeder)
        dfls_list = format_detailed_results(app, region, feeder.devices)

        fault_studies_pd[feeder.obj.loc_name] = [
            study_results_df,
            fdr_open_points,
            dfls_list
        ]

    return fault_studies_pd


def format_study_results(app: pft.Application, devices: List) -> pd.DataFrame:
    """
    Format device summary results for the Summary Results sheet.

    Args:
        app: PowerFactory application instance.
        devices: List of Device dataclasses.

    Returns:
        DataFrame with device fault levels, capacity, and relationships.
    """
    device_list = format_devices()

    for device in devices:
        ds_names = [str(d.obj.loc_name) for d in device.ds_devices]
        us_names = [str(d.obj.loc_name) for d in device.us_devices]
        max_ds_tr = device.max_ds_tr

        device_values = [
            safe_numeric(device.l_l_volts),
            safe_numeric(device.phases),
            safe_numeric(device.ds_capacity),
            safe_numeric(device.max_fl_3ph),
            safe_numeric(device.max_fl_2ph),
            safe_numeric(device.max_fl_pg),
            safe_numeric(device.min_fl_3ph),
            safe_numeric(device.min_fl_2ph),
            safe_numeric(device.min_fl_pg),
            safe_numeric(device.min_sn_fl_2ph),
            safe_numeric(device.min_sn_fl_pg),
            (
                str(max_ds_tr.term.cpSubstat.loc_name)
                if max_ds_tr.term is not None
                else ''
            ),
            safe_numeric(max_ds_tr.load_kva),
            safe_numeric(max_ds_tr.max_ph),
            safe_numeric(max_ds_tr.max_pg),
            ', '.join(ds_names),
            ', '.join(us_names),
        ]

        device_list[str(device.obj.loc_name)] = device_values

    formatted_dev_pd = pd.DataFrame.from_dict(device_list)
    study_results_df = formatted_dev_pd.fillna("")

    return study_results_df


def format_fdr_open_points(feeder) -> pd.DataFrame:
    """
    Format feeder open points for output.

    Args:
        feeder: Feeder dataclass with open_points attribute.

    Returns:
        DataFrame with single column listing open point names.
    """
    open_points = feeder.open_points
    safe_open_points = [str(op.loc_name) for op in open_points]

    fdr_open_points = {'Feeder Open Points': safe_open_points}
    formatted_df = pd.DataFrame.from_dict(fdr_open_points)
    formatted_df = formatted_df.fillna("")

    return formatted_df


def format_detailed_results(
    app: pft.Application,
    region: str,
    devices: List
) -> List[pd.DataFrame]:
    """
    Format detailed terminal-level results for each device.

    Creates a DataFrame for each device containing fault levels and
    reach factors at each terminal in the protection section.

    Args:
        app: PowerFactory application instance.
        region: Network region string.
        devices: List of Device dataclasses.

    Returns:
        List of DataFrames, one per device.
    """
    dfls_list = []

    for device in devices:
        device_name = str(device.obj.loc_name)
        elements = device.sect_terms

        # Calculate reach factors
        dev_reach_factors = device_reach_factors(region, device, elements)

        df = pd.DataFrame({
            'Tfmr Size (kVA)': [None] * len(elements),
            device_name: [e.obj.loc_name for e in elements],
            'Construction': [e.constr for e in elements],
            'Max 3P fault': [safe_numeric(e.max_fl_3ph) for e in elements],
            'Max 2P fault': [safe_numeric(e.max_fl_2ph) for e in elements],
            'Max PG fault': [safe_numeric(e.max_fl_pg) for e in elements],
            'Min 3P fault': [safe_numeric(e.min_fl_3ph) for e in elements],
            'Min 2P fault': [safe_numeric(e.min_fl_2ph) for e in elements],
            'Min PG fault': [safe_numeric(e.min_fl_pg) for e in elements],
            'Min SN 2P fault': [safe_numeric(e.min_sn_fl_2ph) for e in elements],
            'Min SN PG fault': [safe_numeric(e.min_sn_fl_pg) for e in elements],
            # Pickups
            'EF PRI PU': dev_reach_factors.get('ef_pickup', []),
            'EF BU PU': dev_reach_factors.get('bu_ef_pickup', []),
            'PH PRI PU': dev_reach_factors.get('ph_pickup', []),
            'PH BU PU': dev_reach_factors.get('bu_ph_pickup', []),
            'NPS PRI PU': dev_reach_factors.get('nps_pickup', []),
            'NPS BU PU': dev_reach_factors.get('bu_nps_pickup', []),
            # Reach Factors
            'EF PRI RF': dev_reach_factors.get('ef_rf', []),
            'EF BU RF': dev_reach_factors.get('bu_ef_rf', []),
            'PH PRI RF': dev_reach_factors.get('ph_rf', []),
            'PH BU RF': dev_reach_factors.get('bu_ph_rf', []),
            'NPS EF PRI RF': dev_reach_factors.get('nps_ef_rf', []),
            'NPS EF BU RF': dev_reach_factors.get('bu_nps_ef_rf', []),
            'NPS PH PRI RF': dev_reach_factors.get('nps_ph_rf', []),
            'NPS PH BU RF': dev_reach_factors.get('bu_nps_ph_rf', [])
        })

        # Sort by Max PG fault descending
        if 'Max PG fault' in df.columns and not df.empty:
            try:
                df_sorted = df.sort_values(by='Max PG fault', ascending=False)
            except (AttributeError, KeyError):
                df_sorted = df
        else:
            df_sorted = df

        # Map transformer sizes
        try:
            tr_dict = {
                str(load.term.loc_name): safe_numeric(load.load_kva)
                for load in device.sect_loads
                if hasattr(load, 'term') and hasattr(load, 'load_kva')
            }
            df_sorted['Tfmr Size (kVA)'] = df_sorted[device_name].map(tr_dict)
        except AttributeError:
            pass

        dfls_list.append(df_sorted)

    return dfls_list


def format_devices() -> Dict[str, List]:
    """
    Create the device data structure template for Summary Results.

    Returns:
        Dictionary with 'Site Name' key containing row labels.
    """
    device_list = {
        'Site Name': [
            'L-L Voltage (kV)',
            'No. Phases',
            'DS Capacity  (kVA)',
            'Max 3Ph FL',
            'Max 2Ph FL',
            'Max PG FL',
            'Min 3Ph FL',
            'Min 2Ph FL',
            'Min PG FL',
            'Min SN 2P FL',
            'Min SN PG FL',
            'Max DS TR (Site name)',
            'Max TR size (kVA)',
            'TR Max Ph ',
            'TR Max PG',
            'Downstream Devices',
            'Back-up Device'
        ]
    }
    return device_list


# =============================================================================
# DATA CLEANING FUNCTIONS
# =============================================================================

def safe_numeric(value: Any) -> Any:
    """
    Safely convert value to numeric, preserving None/NaN.

    Args:
        value: Value to convert.

    Returns:
        Numeric value or None if conversion fails.
    """
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    try:
        return float(value) if value else None
    except (ValueError, TypeError):
        return None


def clean_string_value(value: Any) -> str:
    """
    Clean string values to remove problematic characters.

    Removes control characters and limits length for Excel compatibility.

    Args:
        value: Value to clean.

    Returns:
        Cleaned string suitable for Excel cells.
    """
    if pd.isna(value) or value is None:
        return ''

    value_str = str(value)

    # Remove control characters (ASCII 0-31 except tab, newline, CR)
    value_str = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', value_str)

    # Replace specific problematic characters
    replacements = {
        '\x00': '', '\x01': '', '\x02': '', '\x03': '',
        '\x04': '', '\x05': '', '\x06': '', '\x07': '',
        '\x08': '', '\x0B': ' ', '\x0C': ' ', '\x0E': '',
        '\x0F': ''
    }

    for old, new in replacements.items():
        value_str = value_str.replace(old, new)

    # Limit to Excel cell character limit
    if len(value_str) > 32767:
        value_str = value_str[:32767]

    return value_str


def safe_set_cell(worksheet, cell_reference: str, value: Any) -> None:
    """
    Safely set cell value with proper cleaning.

    Args:
        worksheet: openpyxl worksheet object.
        cell_reference: Cell reference string (e.g., 'A1').
        value: Value to set.
    """
    try:
        cleaned_value = clean_string_value(value)
        worksheet[cell_reference] = cleaned_value
    except Exception:
        worksheet[cell_reference] = str(value)[:32767] if value else ''


def create_safe_sheet_name(name: str) -> str:
    """
    Create a valid Excel sheet name.

    Excel sheet names have restrictions:
    - Max 31 characters
    - Cannot contain: \\ / ? * [ ] :
    - Cannot be empty
    - Cannot start/end with single quote

    Args:
        name: Proposed sheet name.

    Returns:
        Valid Excel sheet name.
    """
    if not name:
        name = "Sheet"

    name_str = clean_string_value(str(name))

    forbidden_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for char in forbidden_chars:
        name_str = name_str.replace(char, '_')

    name_str = name_str.strip("'")

    if not name_str or name_str.isspace():
        name_str = "Sheet"

    if len(name_str) > 31:
        name_str = name_str[:31]

    name_str = name_str.rstrip()

    return name_str


def fix_string(file_name: str) -> str:
    """
    Remove invalid characters from filename.

    Args:
        file_name: Proposed filename.

    Returns:
        Valid filename with forbidden characters replaced.
    """
    if not file_name:
        return "default_filename"

    file_name_str = str(file_name)
    forbidden_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']

    for char in forbidden_chars:
        file_name_str = file_name_str.replace(char, '_')

    file_name_str = re.sub(r'[\x00-\x1F\x7F]', '_', file_name_str)

    return file_name_str


def ensure_numeric_types(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure numeric values are stored as proper numeric types.

    Converts string representations of numbers to actual numeric types
    for proper Excel formatting.

    Args:
        df: DataFrame to process.

    Returns:
        DataFrame with proper numeric types.
    """
    if df is None or df.empty:
        return df

    df_numeric = df.copy()

    for col in df_numeric.columns:
        col_str = str(col).lower()

        # Skip string columns
        if any(kw in col_str for kw in ['name', 'construction', 'site', 'device']):
            continue

        df_numeric[col] = pd.to_numeric(df_numeric[col], errors='ignore')

        # Convert whole numbers to integers
        if df_numeric[col].dtype in ['float64', 'float32']:
            non_null = df_numeric[col].dropna()
            if len(non_null) > 0:
                all_whole = all(
                    pd.isna(val) or (isinstance(val, (int, float)) and val == int(val))
                    for val in non_null
                )
                if all_whole:
                    df_numeric[col] = df_numeric[col].astype('Int64')

    return df_numeric


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean DataFrame for Excel compatibility.

    Replaces inf values and cleans string columns.

    Args:
        df: DataFrame to clean.

    Returns:
        Cleaned DataFrame.
    """
    if df is None or df.empty:
        return df

    df_clean = df.copy()
    df_clean = df_clean.replace([float('inf'), float('-inf')], None)

    for col in df_clean.columns:
        if df_clean[col].dtype == 'object':
            df_clean[col] = df_clean[col].apply(
                lambda x: clean_string_value(x) if isinstance(x, str) else x
            )

    df_clean.columns = [clean_string_value(str(col)) for col in df_clean.columns]

    return df_clean


# =============================================================================
# COLUMN WIDTH ADJUSTMENT FUNCTIONS
# =============================================================================

def adjust_gen_info_col_size(ws) -> None:
    """Adjust column widths for General Information sheet."""
    ws["A1"].font = Font(bold=True, size=12)
    ws["A2"].font = Font(bold=True, size=12)

    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 18.0

    for cell in ws[27]:
        cell.alignment = Alignment(wrap_text=True)


def adjust_summ_col_size(ws) -> None:
    """Adjust column widths for Summary Results sheet."""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter

        second_row_cell = col[1] if len(col) > 1 else None
        is_tfmr_col = (
            second_row_cell
            and second_row_cell.value
            and str(second_row_cell.value) == "Tfmr Size (kVA)"
        )

        if is_tfmr_col:
            ws.column_dimensions[column].width = 13.71
        else:
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass

            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width


def adjust_detailed_col_size(ws) -> None:
    """Adjust column widths for Detailed Results sheets."""
    ws.row_dimensions[2].height = 30.00

    for cell in ws[2]:
        cell.alignment = Alignment(wrap_text=True)

    device_col_letter = None

    for col in ws.columns:
        column = col[0].column_letter
        second_row_cell = col[1] if len(col) > 1 else None

        is_tfmr_col = (
            second_row_cell
            and second_row_cell.value
            and str(second_row_cell.value) == "Tfmr Size (kVA)"
        )

        if is_tfmr_col:
            ws.column_dimensions[column].width = 13.71
            tr_col_num = second_row_cell.column
            device_col_letter = get_column_letter(tr_col_num + 1)
        elif column == device_col_letter:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass

            ws.column_dimensions[column].width = max_length + 2
        else:
            ws.column_dimensions[column].width = 15


def adjust_cond_damage_col_width(ws) -> None:
    """Adjust column widths for Conductor Damage Results sheets."""
    ws.row_dimensions[1].height = 30.00

    wrap_alignment = Alignment(wrap_text=True)
    for cell in ws[1]:
        cell.alignment = wrap_alignment

    fixed_widths = {
        'E': 13.86, 'F': 17.14, 'G': 14.71, 'H': 16.43,
        'I': 13.57, 'J': 17.14, 'K': 15.71, 'L': 16.29
    }

    for col in ws.columns:
        column_letter = col[0].column_letter

        if column_letter in ['A', 'B', 'C', 'D']:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (AttributeError, TypeError):
                    pass

            ws.column_dimensions[column_letter].width = max_length + 2

        elif column_letter in fixed_widths:
            ws.column_dimensions[column_letter].width = fixed_widths[column_letter]